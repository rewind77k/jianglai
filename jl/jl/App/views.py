# App/views.py
import time
from django.shortcuts import HttpResponse, redirect, render
import os
import json
import requests
from App.models import *
from django.db import transaction, connection
from django.db.models import Q
from io import BytesIO
import base64
from django.utils.deprecation import MiddlewareMixin
from django.http import JsonResponse
from openpyxl import load_workbook
from JiangLai import settings

path = r"F:/test"
os.chdir(path)  # 修改工作路径


# 系统界面
def login(request):
    # if request.method == "POST":
    # return JsonResponse({'code': 200, 'msg': 'success'})
    username = request.GET.get('username')  # 用户名
    password = request.GET.get('password')  # 密码
    if username and password:
        try:
            t_password = Users.objects.get(username=username).password
            t_type = Users.objects.get(username=username).type
            t_name = Users.objects.get(username=username).name
        except:
            return JsonResponse({'code': -4, 'msg': '用户名不存在'})
        else:
            if password == t_password:
                return JsonResponse({'code': 200,
                                     'msg': 'success',
                                     'type': t_type,
                                     'token': username,
                                     'name': t_name})
            else:
                return JsonResponse({'code': -5, 'msg': '密码错误'})
    else:
        return JsonResponse({'code': -3, 'msg': '参数错误'})


def initialize(request):
    wb = load_workbook('initialize.xlsx')
    # 删表实现覆盖操作
    Cotest_Standard.objects.all().delete()
    Teacher.objects.all().delete()
    Student.objects.all().delete()
    Class3_Grade.objects.all().delete()
    Users.objects.all().delete()
    # 生成综测标准表
    sheet1 = wb['cotest_standard']
    for row in sheet1.iter_rows(min_row=2):
        Cotest_Standard.objects.create(
            co_st_id=row[0].value,
            co_st_name=row[1].value,
            co_st_type=row[2].value,
            co_st_verify=bool(row[3].value),
        )
    # 生成教师表
    sheet2 = wb['teacher']
    for row in sheet2.iter_rows(min_row=2):
        Teacher.objects.create(
            te_id=row[0].value,
            te_class=row[1].value,
            te_name=row[2].value,
        )
        Users.objects.create(
            username=row[0].value,
            password='123456',
            name=row[2].value,
            type=row[1].value,
        )
    # 生成学生表
    sheet3 = wb['student']
    for row in sheet3.iter_rows(min_row=2):
        Student.objects.create(
            st_id=row[0].value,
            st_class=row[1].value,
            st_name=row[2].value,
        )
        Users.objects.create(
            username=row[0].value,
            password='123456',
            name=row[2].value,
            type=1,
        )
    # 为每个学生生成初始三级表，图片默认为未上传图片的index
    for student in Student.objects.all():
        for cs in Cotest_Standard.objects.filter(co_st_verify=0):
            Class3_Grade.objects.create(
                st_id=student.st_id,
                co_st_id=cs.co_st_id,
                st_name=student.st_name,
                st_class=student.st_class,
                su_picture='/img/0.jpg',
                cl3_grade=0,
                co_st_verify=0,
            )
    return HttpResponse("操作成功")


def info_list(request):
    data_list = Users.objects.all()
    return render(request, "info_list.html", {"data_list": data_list})


def info_add(request):
    if request.method == 'POST':
        une = request.POST.get('username')
        pwd = request.POST.get('password')
        name = request.POST.get('name')
        ty = request.POST.get('type')
        Users.objects.create(username=une, password=pwd, name=name, type=ty)
        return redirect('/info/list/')
    elif request.method == "GET":
        return render(request, 'info_add.html')
    return HttpResponse("添加失败")


def info_delete(request):
    # if request.method == 'POST':
    nid = request.GET.get('nid')
    Users.objects.filter(id=nid).delete()
    return redirect('/info/list/')


def upload_3(request):
    exc = request.FILES.get('file')
    url = '/exc/'
    old_name = exc.name
    suffix = old_name.rsplit('.')[1]
    up_time = time.strftime('%Y%m%d-%H%M%S', time.localtime(time.time()))
    exc_dir = os.path.join(settings.EXC_ROOT, up_time + '.' + suffix)
    # 文件保存到本地
    destination = open(exc_dir, 'wb+')
    for chunk in exc.chunks():
        destination.write(chunk)
    destination.close()
    # 文件路径以字符串存入数据库
    Supporting_Materials.objects.create(su_ma_index=url + up_time + '.' + suffix, co_st_id=12341, te_id=3)
    # 读取表格处理数据
    wb = load_workbook(exc_dir)
    sheet = wb['政治学习']
    for row in sheet.iter_rows(min_row=2):
        grade = 20
        grade -= (row[3].value + row[4].value + row[5].value) * 5
        if grade < 0:
            grade = 0
        Class3_Grade.objects.create(
            st_class=row[0].value,
            su_picture='',
            cl3_grade=grade,
            st_name=row[2].value,
            co_st_verify=True,
            st_id=row[1].value,
            co_st_id=111,
        )
    sheet = wb['文明修身与遵纪守法']
    for row in sheet.iter_rows(min_row=2):
        grade = 30
        grade -= 2 * row[3].value + 10 * (row[4].value + row[5].value) + 15 * row[6].value + 5 * row[7].value
        if grade < 0:
            grade = 0
        Class3_Grade.objects.create(
            st_class=row[0].value,
            su_picture='',
            cl3_grade=grade,
            st_name=row[2].value,
            co_st_verify=True,
            st_id=row[1].value,
            co_st_id=112,
        )
    sheet = wb['寝室与三级建家']
    for row in sheet.iter_rows(min_row=2):
        grade = 0
        if row[3].value >= 16:
            grade += 10
        elif row[3].value >= 17:
            grade += 12
        elif row[3].value >= 19:
            grade += 15
        grade += row[4].value * 2 - row[5].value
        if grade < 0:
            grade = 0
        if grade > 25:
            grade = 25
        Class3_Grade.objects.create(
            st_class=row[0].value,
            su_picture='',
            cl3_grade=grade,
            st_name=row[2].value,
            co_st_verify=True,
            st_id=row[1].value,
            co_st_id=113,
        )
    sheet = wb['义务献血']
    for row in sheet.iter_rows(min_row=2):
        grade = 0
        if row[5].value == '是':
            grade += 10
        elif row[4].value == '是':
            grade += 1
        Class3_Grade.objects.create(
            st_class=row[0].value,
            su_picture='',
            cl3_grade=grade,
            st_name=row[2].value,
            co_st_verify=True,
            st_id=row[1].value,
            co_st_id=114,
        )
    sheet = wb['智育分数']
    for row in sheet.iter_rows(min_row=2):
        Class3_Grade.objects.create(
            st_class=row[0].value,
            su_picture='',
            cl3_grade=row[4].value,
            st_name=row[2].value,
            co_st_verify=True,
            st_id=row[1].value,
            co_st_id=211,
        )
        Class3_Grade.objects.create(
            st_class=row[0].value,
            su_picture='',
            cl3_grade=row[3].value * 0.7,
            st_name=row[2].value,
            co_st_verify=True,
            st_id=row[1].value,
            co_st_id=311,
        )
    sheet = wb['升旗早操']
    for row in sheet.iter_rows(min_row=2):
        grade = 30 - row[3].value * 10
        Class3_Grade.objects.create(
            st_class=row[0].value,
            su_picture='',
            cl3_grade=grade,
            st_name=row[2].value,
            co_st_verify=True,
            st_id=row[1].value,
            co_st_id=312,
        )
    sheet = wb['组织管理能力']
    for row in sheet.iter_rows(min_row=2):
        Class3_Grade.objects.create(
            st_class=row[0].value,
            su_picture='',
            cl3_grade=row[3].value,
            st_name=row[2].value,
            co_st_verify=True,
            st_id=row[1].value,
            co_st_id=411,
        )
    sheet = wb['校园文化活动']
    for row in sheet.iter_rows(min_row=2):
        Class3_Grade.objects.create(
            st_class=row[0].value,
            su_picture='',
            cl3_grade=row[3].value + row[4].value,
            st_name=row[2].value,
            co_st_verify=True,
            st_id=row[1].value,
            co_st_id=412,
        )
    return JsonResponse({'code': 200, 'msg': '上传成功'})


def upload_21(request):
    exc = request.FILES.get('file')
    url = '/exc/'
    old_name = exc.name
    suffix = old_name.rsplit('.')[1]
    up_time = time.strftime('%Y%m%d-%H%M%S', time.localtime(time.time()))
    exc_dir = os.path.join(settings.EXC_ROOT, up_time + '.' + suffix)
    # 文件保存到本地
    destination = open(exc_dir, 'wb+')
    for chunk in exc.chunks():
        destination.write(chunk)
    destination.close()
    # 文件路径以字符串存入数据库
    Supporting_Materials.objects.create(su_ma_index=url + up_time + '.' + suffix, co_st_id=212, te_id=21)
    # 读取表格处理数据
    wb = load_workbook(exc_dir)
    for i in range(0, 2):
        sheet = wb.worksheets[i]
        for row in sheet.iter_rows(min_row=2):
            res = Class3_Grade.objects.filter(st_id=row[1].value)
            exists = res.filter(co_st_id='212').exists()
            if exists:
                student = res.objects.get(co_st_id=212)
                student.cl3_grade = max(student.cl3_grade, 5 + (ord('C') - ord(row[3].value)) * 2 - row[4].value)
                student.save()
            else:
                Class3_Grade.objects.create(
                    st_class=row[0].value,
                    su_picture='',
                    cl3_grade=5 + (ord('C') - ord(row[3].value)) * 2 - row[4].value,
                    st_name=row[2].value,
                    co_st_verify=True,
                    st_id=row[1].value,
                    co_st_id='212',
                )
            print(5 + (ord('C') - ord(row[3].value)) * 2 - row[4].value)
    return JsonResponse({'code': 200, 'msg': '上传成功'})


def upload_1(request):
    st_id = request.META.get('HTTP_AUTHORIZATION')
    co_st_id = request.GET.get('co_st_id')
    img = request.FILES.get('file')
    url = '/img/'
    old_name = img.name
    suffix = old_name.rsplit('.')[1]
    up_time = time.strftime('%Y%m%d-%H%M%S', time.localtime(time.time()))
    img_dir = os.path.join(settings.IMG_ROOT, up_time + '.' + suffix)
    destination = open(img_dir, 'wb+')
    for chunk in img.chunks():
        destination.write(chunk)
    destination.close()
    student = Class3_Grade.objects.filter(st_id=st_id).get(co_st_id=co_st_id)
    student.su_picture = url + up_time + '.' + suffix
    student.save()
    return JsonResponse({'code': 200, 'msg': '上传成功', 'st_id': st_id, 'co_st_id': co_st_id})


def grade_calculate(request):
    # 获取学生的queryset
    students = Student.objects.all()
    # 处理每个学生的成绩
    Class2_Grade.objects.all().delete()
    Class1_Grade.objects.all().delete()
    for student in students:
        res = Class3_Grade.objects.filter(st_id=student.st_id)
        grade_2 = [0, 0, 0, 0, 0]
        grade_race = 0
        for item in res:
            # 主流和小众竞赛取其最大值
            if item.co_st_id == '204' or item.co_st_id == '212':
                grade_race = max(grade_race, item.cl3_grade)
            else:
                grade_2[int(item.co_st_id[0])] += item.cl3_grade
            grade_2[2] += grade_race
        # 处理溢出分数
        for i in range(1, 5):
            if grade_2[i] > 100:
                grade_2[i] = 100
            Class2_Grade.objects.create(
                st_id=res[0].st_id,
                co_st_type=i,
                st_name=res[0].st_name,
                st_class=res[0].st_class,
                cl2_grade=grade_2[i]
            )
        grade_1 = grade_2[1] * 0.2 + grade_2[2] * 0.7 + grade_2[3] * 0.05 + grade_2[4] * 0.05
        Class1_Grade.objects.create(
            st_id=res[0].st_id,
            st_name=res[0].st_name,
            st_class=res[0].st_class,
            cl1_grade=grade_1
        )
    # 对每个学生成绩进行排序: 2%为一等奖,10%为二等奖,20%为三等奖
    class_map = {}  # 记录所有班级
    for student in Student.objects.all():
        class_map[student.st_class] = 1

    for key, value in class_map.items():
        res = Class1_Grade.objects.filter(st_class=key).order_by('-cl1_grade')
        length = len(res)
        prize_1 = int(length * 0.02)
        if prize_1 < 1:
            prize_1 = 1
        prize_2 = int(length * 0.12)
        prize_3 = int(length * 0.32)
        cnt = 0
        for item in res:
            if cnt < prize_1:
                item.cl1_class = '一等奖'
            elif cnt < prize_2:
                item.cl1_class = '二等奖'
            elif cnt < prize_3:
                item.cl1_class = '三等奖'
            else:
                item.cl1_class = ' '
            cnt += 1
            item.save()
    return JsonResponse({'code':200, 'msg': '操作成功'})


def counselor_verify(request):
    op = request.GET.get('op')
    # 返回初始界面表单
    if op == '1':
        student_list = []
        students = Student.objects.all()
        for student in students:
            # 默认审核到401项目结束审核
            flag = Class3_Grade.objects.filter(st_id=student.st_id).get(co_st_id='401').co_st_verify
            verify_info = '审核完成' if flag else '未审核'
            student_list.append({'id': student.st_id,
                                 'class': student.st_class,
                                 'name': student.st_name,
                                 'state': verify_info})
        return JsonResponse({'code': 200, 'msg': '返回表单成功', 'data': student_list})
    # 返回弹框界面信息
    elif op == '2':
        photo_list = []
        st_id = request.GET.get('st_id')
        res = Class3_Grade.objects.filter(st_id=st_id)
        for item in res:
            index = "../files" + item.su_picture
            photo_list.append({
                'id': item.co_st_id,
                'name': Cotest_Standard.objects.get(co_st_id=item.co_st_id).co_st_name,
                'index': index,
                'grade': item.cl3_grade,
            })
        return JsonResponse({'code': 200, 'msg': '返回图片成功', 'data': photo_list})
    # 读取第三次交互返回的审核结果，进行数据库修改
    elif op == '3':
        data_list = request.GET.get("info_list")
        st_id = request.GET.get('st_id')
        data_list = data_list.split(',')
        for student in Class3_Grade.objects.filter(st_id=st_id):
            if student.co_st_id == '101':
                student.cl3_grade = data_list[0]
                student.co_st_verify = 1
            if student.co_st_id == '102':
                student.cl3_grade = data_list[1]
                student.co_st_verify = 1
            if student.co_st_id == '201':
                student.cl3_grade = data_list[2]
                student.co_st_verify = 1
            if student.co_st_id == '202':
                student.cl3_grade = data_list[3]
                student.co_st_verify = 1
            if student.co_st_id == '203':
                student.cl3_grade = data_list[4]
                student.co_st_verify = 1
            if student.co_st_id == '204':
                student.cl3_grade = data_list[5]
                student.co_st_verify = 1
            if student.co_st_id == '301':
                student.cl3_grade = data_list[6]
                student.co_st_verify = 1
            if student.co_st_id == '401':
                student.cl3_grade = data_list[7]
                student.co_st_verify = 1
            student.save()
        return JsonResponse({'code': 200, 'msg': '提交成功'})

    # 读取操作码错误
    else:
        return JsonResponse({'code': -3, 'msg': '操作失败'})


def student_lookover(request):
    st_id = request.GET.get('username')
    student = Class1_Grade.objects.get(st_id=st_id)
    dic = {
        'id': student.st_id,
        'class': student.st_class,
        'grade': student.cl1_grade,
        'cl1_class': student.cl1_class,
    }
    return JsonResponse({'code': '200', 'msg': '返回成功', 'data': dic})


def counselor_lookover(request):
    op = request.GET.get('op')
    # 返回所有学生信息表单
    if op == '1':
        student_list = []
        for student in Class1_Grade.objects.all().order_by('st_class','-cl1_grade'):
            student_list.append({
                'st_id': student.st_id,
                'st_class': student.st_class,
                'st_name': student.st_name,
                'cl1_grade': round(student.cl1_grade, 2),
                'cl1_class': student.cl1_class
            })
        return JsonResponse({'code': 200, 'msg': '返回成功', 'data': student_list})
    elif op == '2':
        st_id = request.GET.get('st_id')
        student = Student.objects.get(st_id=st_id)
        student_list = [{
            'st_id': student.st_id,
            'st_class': student.st_class,
            'st_name': student.st_name,
        }]
        grade1 = Class1_Grade.objects.get(st_id=st_id)
        student_list.append({
            'cl1_grade': round(grade1.cl1_grade, 2),
            'cl1_class': grade1.cl1_class,
        })
        grade2 = Class2_Grade.objects.filter(st_id=st_id)
        cl2_list = [
            0,
            round(grade2.get(co_st_type='1').cl2_grade, 2),
            round(grade2.get(co_st_type='2').cl2_grade, 2),
            round(grade2.get(co_st_type='3').cl2_grade, 2),
            round(grade2.get(co_st_type='4').cl2_grade, 2),
        ]
        student_list.append({'cl2_list': cl2_list})
        cl3_dic = {}
        for student in Class3_Grade.objects.filter(st_id=st_id):
            if student.co_st_id == '101':
                cl3_dic['101'] = round(student.cl3_grade, 2)
            if student.co_st_id == '102':
                cl3_dic['102'] = round(student.cl3_grade, 2)
            if student.co_st_id == '201':
                cl3_dic['201'] = round(student.cl3_grade, 2)
            if student.co_st_id == '202':
                cl3_dic['202'] = round(student.cl3_grade, 2)
            if student.co_st_id == '203':
                cl3_dic['203'] = round(student.cl3_grade, 2)
            if student.co_st_id == '204':
                cl3_dic['204'] = round(student.cl3_grade, 2)
            if student.co_st_id == '301':
                cl3_dic['301'] = round(student.cl3_grade, 2)
            if student.co_st_id == '401':
                cl3_dic['401'] = round(student.cl3_grade, 2)
        student_list.append({'cl3_dic': cl3_dic})
        return JsonResponse({'code': 200, 'msg': '返回成功', 'data': student_list})
    else:
        return JsonResponse({'code': '-3', 'msg': '参数错误'})
